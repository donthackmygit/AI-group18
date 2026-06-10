from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = PROJECT_ROOT / "data" / "processed" / "cleaned_text"
OUTPUT_DIR = PROJECT_ROOT / "data" / "processed" / "structured_documents"
LOG_PATH = PROJECT_ROOT / "data" / "processed" / "structure_parser_log.csv"
REGISTRY_PATH = PROJECT_ROOT / "data" / "metadata" / "document_registry.xlsx"


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


CHAPTER_PATTERN = re.compile(
    r"^(Chương|CHƯƠNG)\s+([IVXLCDM]+|\d+)\.?\s*(.*)$",
    re.IGNORECASE,
)
SECTION_PATTERN = re.compile(
    r"^(Mục|MỤC)\s+(\d+[A-Za-z]?)\.?\s*(.*)$",
    re.IGNORECASE,
)
ARTICLE_PATTERN = re.compile(
    r"^(Điều|ĐIỀU)\s+(\d+[a-zA-Z]?)\s*[\.:]?\s*(.*)$",
    re.IGNORECASE,
)
CLAUSE_PATTERN = re.compile(r"^(\d{1,3})\.\s+(.+)$")
POINT_PATTERN = re.compile(r"^([a-zđ])\)\s+(.+)$", re.IGNORECASE)


REGISTRY_METADATA_FIELDS = [
    "document_id",
    "file_name",
    "title",
    "document_number",
    "document_type",
    "issuing_authority",
    "issue_date",
    "effective_date",
    "expiry_date",
    "status",
    "source_type",
    "source_url",
    "local_path",
    "download_date",
    "topics",
    "version",
    "notes",
]


def load_registry_metadata(registry_path: Path) -> dict[str, dict[str, Any]]:
    if not registry_path.exists():
        return {}

    workbook = load_workbook(registry_path, read_only=True, data_only=True)
    sheet = workbook["document_registry"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    indexes = {header: index for index, header in enumerate(headers or [])}

    metadata_by_id: dict[str, dict[str, Any]] = {}
    for row in rows:
        if not row or not any(value is not None and str(value).strip() for value in row):
            continue

        document_id_index = indexes.get("document_id")
        if document_id_index is None:
            continue

        document_id = str(row[document_id_index] or "").strip()
        if not document_id:
            continue

        metadata: dict[str, Any] = {}
        for field in REGISTRY_METADATA_FIELDS:
            index = indexes.get(field)
            value = row[index] if index is not None and index < len(row) else None
            metadata[field] = "" if value is None else value
        metadata_by_id[document_id] = metadata

    return metadata_by_id


def normalize_content(lines: list[str]) -> str:
    return "\n".join(line for line in lines if line is not None).strip()


def finalize_article(article: dict[str, Any] | None) -> dict[str, Any] | None:
    if article is None:
        return None

    article["content"] = normalize_content(article.pop("_content_lines", []))
    for clause in article["clauses"]:
        clause["content"] = normalize_content(clause.pop("_content_lines", []))
        for point in clause["points"]:
            point["content"] = normalize_content(point.pop("_content_lines", []))
    return article


def paragraphs_from_text(text: str) -> list[dict[str, Any]]:
    paragraphs = []
    for index, line in enumerate(text.splitlines(), start=1):
        line = line.strip()
        if line:
            paragraphs.append({"paragraph_id": index, "content": line})
    return paragraphs


def is_official_dispatch(document_id: str, metadata: dict[str, Any] | None) -> bool:
    document_type = str((metadata or {}).get("document_type", "")).strip().lower()
    return document_type == "công văn" or document_id.startswith("DISPATCH_")


def parse_document_structure(
    document_id: str,
    text: str,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not text.strip():
        return {
            "document_id": document_id,
            "metadata": metadata or {},
            "parsed_at": datetime.now().isoformat(timespec="seconds"),
            "parse_status": "EMPTY",
            "article_count": 0,
            "clause_count": 0,
            "point_count": 0,
            "paragraph_count": 0,
            "preamble": "",
            "articles": [],
            "paragraphs": [],
        }

    if is_official_dispatch(document_id, metadata):
        paragraphs = paragraphs_from_text(text)
        return {
            "document_id": document_id,
            "metadata": metadata or {},
            "parsed_at": datetime.now().isoformat(timespec="seconds"),
            "parse_status": "NO_ARTICLE_FOUND",
            "article_count": 0,
            "clause_count": 0,
            "point_count": 0,
            "paragraph_count": len(paragraphs),
            "preamble": "",
            "articles": [],
            "paragraphs": paragraphs,
        }

    lines = text.splitlines()

    current_chapter: str | None = None
    current_section: str | None = None
    current_article: dict[str, Any] | None = None
    current_clause: dict[str, Any] | None = None
    current_point: dict[str, Any] | None = None

    articles: list[dict[str, Any]] = []
    preamble_lines: list[str] = []

    def flush_article() -> None:
        nonlocal current_article, current_clause, current_point
        finalized = finalize_article(current_article)
        if finalized is not None:
            articles.append(finalized)
        current_article = None
        current_clause = None
        current_point = None

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue

        chapter_match = CHAPTER_PATTERN.match(line)
        if chapter_match:
            flush_article()
            current_chapter = line
            current_section = None
            continue

        section_match = SECTION_PATTERN.match(line)
        if section_match:
            flush_article()
            current_section = line
            continue

        article_match = ARTICLE_PATTERN.match(line)
        if article_match:
            flush_article()
            article_number = article_match.group(2)
            article_title = article_match.group(3).strip()
            current_article = {
                "article": f"Điều {article_number}",
                "article_number": article_number,
                "title": article_title,
                "chapter": current_chapter,
                "section": current_section,
                "_content_lines": [line],
                "clauses": [],
            }
            current_clause = None
            current_point = None
            continue

        if current_article is None:
            preamble_lines.append(line)
            continue

        current_article["_content_lines"].append(line)

        clause_match = CLAUSE_PATTERN.match(line)
        if clause_match:
            current_clause = {
                "clause": clause_match.group(1),
                "_content_lines": [line],
                "points": [],
            }
            current_article["clauses"].append(current_clause)
            current_point = None
            continue

        point_match = POINT_PATTERN.match(line)
        if point_match and current_clause is not None:
            current_point = {
                "point": point_match.group(1),
                "_content_lines": [line],
            }
            current_clause["points"].append(current_point)
            current_clause["_content_lines"].append(line)
            continue

        if current_point is not None:
            current_point["_content_lines"].append(line)
        if current_clause is not None:
            current_clause["_content_lines"].append(line)

    flush_article()

    clause_count = sum(len(article["clauses"]) for article in articles)
    point_count = sum(
        len(clause["points"])
        for article in articles
        for clause in article["clauses"]
    )
    paragraphs = [] if articles else paragraphs_from_text(text)

    if articles:
        parse_status = "OK"
    else:
        parse_status = "NO_ARTICLE_FOUND"

    return {
        "document_id": document_id,
        "metadata": metadata or {},
        "parsed_at": datetime.now().isoformat(timespec="seconds"),
        "parse_status": parse_status,
        "article_count": len(articles),
        "clause_count": clause_count,
        "point_count": point_count,
        "paragraph_count": len(paragraphs),
        "preamble": normalize_content(preamble_lines),
        "articles": articles,
        "paragraphs": paragraphs,
    }


def write_log(log_rows: list[dict[str, Any]], log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "document_id",
        "input_path",
        "output_path",
        "article_count",
        "clause_count",
        "point_count",
        "paragraph_count",
        "status",
        "note",
    ]
    with log_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(log_rows)


def run_parser(
    input_dir: Path,
    output_dir: Path,
    log_path: Path,
    registry_path: Path,
) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    registry = load_registry_metadata(registry_path)

    input_files = sorted(
        file for file in input_dir.glob("*.txt") if file.name != ".gitkeep"
    )
    log_rows: list[dict[str, Any]] = []
    counters = {"success": 0, "warning": 0, "error": 0}

    for input_file in input_files:
        document_id = input_file.stem
        output_file = output_dir / f"{document_id}.json"
        try:
            text = input_file.read_text(encoding="utf-8", errors="ignore")
            parsed = parse_document_structure(
                document_id=document_id,
                text=text,
                metadata=registry.get(document_id, {}),
            )
            output_file.write_text(
                json.dumps(parsed, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            status = parsed["parse_status"]
            if status == "OK":
                counters["success"] += 1
                print(
                    f"[SUCCESS] {input_file.name}: "
                    f"{parsed['article_count']} articles, "
                    f"{parsed['clause_count']} clauses, "
                    f"{parsed['point_count']} points"
                )
            else:
                counters["warning"] += 1
                if status == "NO_ARTICLE_FOUND" and parsed["paragraph_count"]:
                    print(
                        f"[WARNING] {input_file.name}: {status}, "
                        f"{parsed['paragraph_count']} paragraphs"
                    )
                else:
                    print(f"[WARNING] {input_file.name}: {status}")

            log_rows.append(
                {
                    "document_id": document_id,
                    "input_path": str(input_file.relative_to(PROJECT_ROOT)),
                    "output_path": str(output_file.relative_to(PROJECT_ROOT)),
                    "article_count": parsed["article_count"],
                    "clause_count": parsed["clause_count"],
                    "point_count": parsed["point_count"],
                    "paragraph_count": parsed["paragraph_count"],
                    "status": status,
                    "note": (
                        "official_dispatch_paragraphs_only"
                        if is_official_dispatch(document_id, registry.get(document_id, {}))
                        and status == "NO_ARTICLE_FOUND"
                        else ""
                    ),
                }
            )
        except Exception as exc:
            counters["error"] += 1
            print(f"[ERROR] {input_file.name}: {exc}")
            log_rows.append(
                {
                    "document_id": document_id,
                    "input_path": str(input_file.relative_to(PROJECT_ROOT)),
                    "output_path": "",
                    "article_count": 0,
                    "clause_count": 0,
                    "point_count": 0,
                    "paragraph_count": 0,
                    "status": "ERROR",
                    "note": str(exc),
                }
            )

    write_log(log_rows, log_path)

    print("\n===== STRUCTURE PARSER SUMMARY =====")
    print(f"Success: {counters['success']}")
    print(f"Warning: {counters['warning']}")
    print(f"Error:   {counters['error']}")
    print(f"Output:  {output_dir}")
    print(f"Log:     {log_path}")

    return 1 if counters["error"] else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse legal document structure from cleaned text files."
    )
    parser.add_argument("--input-dir", type=Path, default=INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--log-path", type=Path, default=LOG_PATH)
    parser.add_argument("--registry", type=Path, default=REGISTRY_PATH)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return run_parser(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        log_path=args.log_path,
        registry_path=args.registry,
    )


if __name__ == "__main__":
    sys.exit(main())
